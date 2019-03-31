
from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def get_custom_ttable(custom: str) -> None:
	'''Creates a new worksheet whith custom timetable
	:param custom: faculty name, room number, course name/id
	'''


	read_wb = load_workbook('Timetable.xlsx')
	wb = Workbook()
	for sheet in read_wb.worksheets:
		read_ws = sheet
		write_ws = wb.create_sheet(sheet.title)

		###This writes the contents of the original timetable to new timetable
		for row in range(1,61):
			for col in range(1,13):
				write_ws.cell(row, col, read_ws.cell(row, col).value)

		for row in range(11,61):
			for col in range(3,13):
				cell = str(write_ws.cell(row, col).value)
				if custom not in cell and cell != '':
					write_ws.cell(row, col, '')

	wb.save(custom + '.xlsx')	

	# ws = workbook.copy_worksheet(workbook['W-1'])
	# ws.title = custom


get_custom_ttable('Ramzan Ali')
