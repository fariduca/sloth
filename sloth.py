from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from unMergeExcelCell import unMergeExcelCell
from cnv_xls_to_xlsx import cvt_xls_to_xlsx


def convert_to_list(worksheet) -> list:
	'''Writing the contents of the chopped TT into a list of form: 
	[[date,subject block 1,subject block 2,...][date, subject block 1,...]...]
	'''
	ls = []
	
	lishnie = ('Free Time', 'Dinner', 'Faculty Meeting')
	for col in range(1, 6):
		ls.append([worksheet.cell(1, col).value])    
		s = set()
		for row in range(2, 50):
			if worksheet.cell(row, col).value and worksheet.cell(row, col).value not in lishnie:
				s.add(worksheet.cell(row, col).value)
		
		for i in s:
			ls[col-1].append(i)

	return ls

def create_xlsx_for_import(workbook: Workbook) -> None:
    '''Creates a new exel file fomatted for import to Outlook calendar'''  
   
    wb = workbook

    new_wb = load_workbook('template.xlsx')    #We're using a ready empty exel file as a template
    ws = new_wb.active
    
    ws['A1'] = 'Start Date'
    ws['B1'] = 'Subject'
    ws['C1'] = 'CourseID'
    ws['D1'] = 'Description'
    ws['E1'] = 'Start Time'
    ws['F1'] = 'End Time'
    ws['G1'] = 'Location'
    ws['H1'] = 'Reminder on/off'
    ws['I1'] = 'Reminder Time'
    ws['J1'] = 'Reminder Date'
    

    ###Sets the coulumn width to 20 of the first seven columns
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    to_row = 2      #This variable keeps track of the row we're writing into

    for worksheet in wb.worksheets:    
        ###Writing the contents of the chopped TT into a list of form: 
        # [[date,subject block 1,subject block 2,...][date, subject block 1,...]...]
        ls = convert_to_list(worksheet)
       
		###This method checks whether the entities are written correctly
        check_TT_entities(ls)
        
        for i in range(len(ls)):
            for k in range(1, len(ls[i])):
                
                ws.cell(to_row, 1, ls[i][0])
                
                title_ls = ls[i][k].split('\n')
                #Gets rid of the spaces near the strings
                for j in range(len(title_ls)):
                    title_ls[j] = title_ls[j].strip()
                
                #Separating the time into start and end
                temp_ls = title_ls[3].split('-')
                title_ls.pop(3)
                
                title_ls.insert(3, temp_ls[1])
                title_ls.insert(3, temp_ls[0])
                for g in range(len(title_ls)):
                    title_ls[g] = title_ls[g].strip('(')
                    title_ls[g] = title_ls[g].strip(')')
                
                for z in range(len(title_ls)):
                    ws.cell(to_row, z+2, title_ls[z])

                #Setting Reminder settings
                ws.cell(to_row, 8, 'On')
                formula_1 = "=E"+str(to_row)+"-TIME(0,5,0)"
                ws.cell(to_row, 9, formula_1)
                ws.cell(to_row, 10, ls[i][0])

                to_row += 1
                
    new_wb.save('file_for_import.xlsx')
    


def clear_TT(cohort: str) -> Workbook:

    ###Unmerge cells and fill them with content. An xls file is created, thus we need to convert it to xlsx
    wb_name = unMergeExcelCell('Timetable.xlsx')

    ###Converting xls to xlsx
    wb = cvt_xls_to_xlsx(wb_name)

    #wb = load_workbook(wb_name)
    #write_wb = Workbook()

    for sheet in wb.sheetnames:    
        ws = wb.get_sheet_by_name(sheet)
        #write_ws = write_wb.create_sheet(sheet)

        ###Copying the contents of Timetable into new workbook
        # for row in range(1,61):
        #     for col in range(1,13):
        #         write_ws.cell(row, col, ws.cell(row, col).value)
        #         write_ws.cell(row, col).alignment = Alignment(wrap_text=True)

        ###Sets the coulumn width to 20 of the first seven columns
        # for i in range(1, 11):
        #     write_ws.column_dimensions[get_column_letter(i)].width = 20


        # write_ws.delete_rows(1, 2)
        # write_ws.delete_cols(1, 2)
        # write_ws.delete_rows(2, 7)
        ws.delete_rows(1, 2)
        ws.delete_cols(1, 2)
        ws.delete_rows(2, 7)
        
        if cohort == 'cohort1':
            for i in range(2,7):
                ws.delete_cols(i)

        elif cohort == 'cohort2':
            for i in range(1,6):
                ws.delete_cols(i)
        
        ws.delete_rows(55, 10)

    wb.save("chopped_"+cohort+".xlsx")
    return wb


def check_TT_entities(ls: list):
    
    for i in range(len(ls)):
        for k in range(1, len(ls[i])):
            title_ls = ls[i][k].split('\n')
            if len(title_ls) != 5:
                raise Exception('Worng syntax!!!\nMistake found. Date: ' + str(ls[i][0]) + '\nBLOCK: '+ str(title_ls) + '\nThe mistake is '+
						'in how some of the courses are recorded. Please correct the syntax!!!\nThe format you want '+
                        'to follow is: \n\tCourse name \n\tCourse ID \n\tInstructor \n\tTime \n\tLocation/Room')



head_wb = clear_TT('cohort1')
clear_TT('cohort2')

create_xlsx_for_import(head_wb)

